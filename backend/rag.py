"""
RAG 파이프라인.

문서 인입(ingest): 파싱 → 청킹 → 임베딩(Ollama) → 인덱스 저장(per-doc JSON)
질의(answer):     질문 임베딩 → 코사인 검색 → 컨텍스트 구성 → LLM 생성 → 인용 반환

임베딩/생성은 Ollama 를 사용하므로 실행 시 Ollama + 해당 모델이 필요하다.
벡터 검색은 별도 벡터DB·검색엔진 없이 JSON 인덱스 + numpy 코사인 유사도로 구현한다.
로드한 인덱스는 파일 mtime 기준으로 메모리에 캐시하여 질의마다 재파싱하지 않는다.
"""

import json
import os
import re

import numpy as np
import requests

from backend.ollama import OLLAMA_HOST
from backend.paths import index_dir

EMBED_MODEL = "bge-m3"          # RAG 임베딩 기본 모델
CHUNK_SIZE = 1000               # 청크 문자 수
CHUNK_OVERLAP = 150
TOP_K = 6                       # 검색 청크 수
SCORE_THRESHOLD = 0.4           # 코사인 유사도 하한 — 미만 청크는 무관한 것으로 보고 문맥/출처에서 제외
# 생성 컨텍스트 길이. Ollama 기본값(4096)은 RAG 프롬프트(문맥 6청크 ≈ 3000토큰)가 거의
# 다 차지해 답변/추론 공간이 부족 → done_reason='length' 로 잘리거나 빈 응답이 난다. 넉넉히 늘린다.
NUM_CTX = 8192


# ---------- 파싱 ----------
def parse_file(path: str) -> list[dict]:
    """파일을 (page, text) 단위 블록 리스트로 파싱."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        return _parse_pdf(path)
    if ext == ".docx":
        return _parse_docx(path)
    if ext == ".xlsx":
        return _parse_xlsx(path)
    if ext == ".hwpx":
        return _parse_hwpx(path)
    if ext in (".txt", ".md"):
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return [{"page": None, "text": f.read()}]
    raise ValueError(f"지원하지 않는 형식입니다: {ext}")


def _parse_pdf(path: str) -> list[dict]:
    from pypdf import PdfReader
    reader = PdfReader(path)
    blocks = []
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if text.strip():
            blocks.append({"page": i, "text": text})
    return blocks


def _parse_docx(path: str) -> list[dict]:
    import docx
    doc = docx.Document(path)
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return [{"page": None, "text": text}]


def _parse_xlsx(path: str) -> list[dict]:
    """
    XLSX(엑셀) 파싱. 시트마다 한 블록으로, 각 행의 셀을 탭으로 이어 텍스트화한다.
    수식 셀은 마지막 저장 시 캐시된 계산값(data_only=True)을 사용한다. 빈 행은 건너뛴다.
    각 블록 앞에 "[시트: 이름]" 헤더를 붙여 검색·인용 시 어느 시트인지 드러나게 한다.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    blocks = []
    try:
        for ws in wb.worksheets:
            lines = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                if any(c.strip() for c in cells):
                    lines.append("\t".join(cells).rstrip())
            text = "\n".join(lines).strip()
            if text:
                blocks.append({"page": None, "text": f"[시트: {ws.title}]\n{text}"})
    finally:
        wb.close()
    if not blocks:
        raise ValueError("XLSX 에서 텍스트를 추출하지 못했습니다.")
    return blocks


def _parse_hwpx(path: str) -> list[dict]:
    """
    HWPX(한글, OWPML) 파싱. ZIP+XML 구조라 표준 라이브러리만으로 텍스트를 추출한다.
    Contents/section*.xml 의 문단(<hp:p>) 단위로, 같은 문단 안 텍스트 런(<hp:t>)은
    공백 없이 이어붙이고(런 경계에서 단어가 갈리는 것 방지) 문단 사이는 개행으로 구분.
    표 셀 안 문단도 동일하게 수집된다. (DOCX 파서와 같은 '평문 텍스트' 수준)
    """
    import zipfile
    from xml.etree import ElementTree as ET

    parts: list[str] = []
    with zipfile.ZipFile(path) as z:
        sections = sorted(n for n in z.namelist()
                          if n.startswith("Contents/section") and n.endswith(".xml"))
        for name in sections:
            try:
                root = ET.fromstring(z.read(name))
            except ET.ParseError:
                continue
            for el in root.iter():
                tag = el.tag.split("}")[-1]    # 네임스페이스 제거
                if tag == "p":
                    parts.append("\n")          # 문단 경계
                elif tag == "t" and el.text:
                    parts.append(el.text)       # 문단 내 런은 공백 없이 연결
    text = "".join(parts).strip()
    if not text:
        raise ValueError("HWPX 에서 텍스트를 추출하지 못했습니다.")
    return [{"page": None, "text": text}]


# ---------- 청킹 ----------
def chunk_blocks(blocks: list[dict]) -> list[dict]:
    chunks = []
    for blk in blocks:
        text = " ".join(blk["text"].split())  # 공백 정규화
        page = blk["page"]
        start = 0
        while start < len(text):
            piece = text[start:start + CHUNK_SIZE]
            if piece.strip():
                chunks.append({"text": piece, "page": page})
            start += CHUNK_SIZE - CHUNK_OVERLAP
    return chunks


# ---------- 임베딩 (Ollama) ----------
def embed_texts(texts: list[str], model: str = EMBED_MODEL) -> list[list[float]]:
    """배치 임베딩. /api/embed 우선, 실패 시 /api/embeddings 폴백."""
    try:
        r = requests.post(
            f"{OLLAMA_HOST}/api/embed",
            json={"model": model, "input": texts},
            timeout=120,
        )
        r.raise_for_status()
        data = r.json()
        if "embeddings" in data:
            return data["embeddings"]
    except requests.RequestException:
        pass
    # 폴백: 단건 반복
    out = []
    for t in texts:
        r = requests.post(
            f"{OLLAMA_HOST}/api/embeddings",
            json={"model": model, "prompt": t},
            timeout=120,
        )
        r.raise_for_status()
        out.append(r.json()["embedding"])
    return out


# ---------- 인덱스 저장/로드 ----------
def _index_path(doc_id: str) -> str:
    return os.path.join(index_dir(), f"{doc_id}.json")


def save_index(doc_id: str, doc_name: str, chunks: list[dict]) -> None:
    with open(_index_path(doc_id), "w", encoding="utf-8") as f:
        json.dump({"doc_id": doc_id, "doc_name": doc_name, "chunks": chunks},
                  f, ensure_ascii=False)
    _index_cache.pop(doc_id, None)  # 내용이 바뀌었으니 캐시 무효화


def load_index(doc_id: str) -> dict | None:
    path = _index_path(doc_id)
    if not os.path.isfile(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def delete_index(doc_id: str) -> None:
    path = _index_path(doc_id)
    if os.path.isfile(path):
        os.remove(path)
    _index_cache.pop(doc_id, None)


# ---------- 검색용 인덱스 캐시 ----------
# doc_id -> {mtime, doc_name, metas, matrix(np.ndarray), norms(np.ndarray)}
# 벡터를 numpy 행렬로 미리 쌓아 두어 질의 시 벡터화 연산으로 코사인을 계산한다.
_index_cache: dict[str, dict] = {}


def _load_for_search(doc_id: str) -> dict | None:
    """검색용 인덱스 항목을 반환. 파일 mtime 이 캐시와 다르면 다시 로드한다."""
    path = _index_path(doc_id)
    if not os.path.isfile(path):
        _index_cache.pop(doc_id, None)
        return None
    mtime = os.path.getmtime(path)
    cached = _index_cache.get(doc_id)
    if cached and cached["mtime"] == mtime:
        return cached
    idx = load_index(doc_id)
    if not idx:
        _index_cache.pop(doc_id, None)
        return None
    metas = [c for c in idx.get("chunks", []) if c.get("vector")]
    if metas:
        matrix = np.asarray([c["vector"] for c in metas], dtype=np.float64)
        norms = np.linalg.norm(matrix, axis=1)
    else:
        matrix = np.zeros((0, 0), dtype=np.float64)
        norms = np.zeros((0,), dtype=np.float64)
    entry = {"mtime": mtime, "doc_name": idx["doc_name"],
             "metas": metas, "matrix": matrix, "norms": norms}
    _index_cache[doc_id] = entry
    return entry


# ---------- 인입 ----------
def ingest(doc_id: str, doc_name: str, path: str,
           embed_model: str = EMBED_MODEL) -> int:
    """문서를 파싱·청킹·임베딩하고 인덱스로 저장. 청크 수 반환."""
    blocks = parse_file(path)
    chunks = chunk_blocks(blocks)
    if not chunks:
        raise ValueError("문서에서 텍스트를 추출하지 못했습니다.")
    vectors = embed_texts([c["text"] for c in chunks], model=embed_model)
    for c, v in zip(chunks, vectors):
        c["vector"] = v
    save_index(doc_id, doc_name, chunks)
    return len(chunks)


# ---------- 검색 ----------
def search(query: str, doc_ids: list[str], top_k: int = TOP_K,
           embed_model: str = EMBED_MODEL) -> list[dict]:
    """주어진 문서들에서 질문과 유사한 청크 top_k 를 반환."""
    qvec = np.asarray(embed_texts([query], model=embed_model)[0], dtype=np.float64)
    qnorm = float(np.linalg.norm(qvec))
    scored = []
    for did in doc_ids:
        idx = _load_for_search(did)
        if not idx or idx["matrix"].shape[0] == 0:
            continue
        # 코사인 = (M·q) / (|M| · |q|). 분모 0 인 행은 0.0 (기존 _cosine 과 동일).
        dots = idx["matrix"] @ qvec
        denom = idx["norms"] * qnorm
        sims = np.divide(dots, denom, out=np.zeros_like(dots), where=denom > 0)
        for c, s in zip(idx["metas"], sims):
            scored.append({
                "score": float(s),
                "text": c["text"],
                "page": c.get("page"),
                "doc_name": idx["doc_name"],
                "doc_id": did,
            })
    scored.sort(key=lambda x: x["score"], reverse=True)
    top = scored[:top_k]
    # 유사도 하한선 적용: 임계값 이상만 사용한다. 전부 미달이면 최상위 1개만 남겨
    # 문맥이 완전히 비지 않게 하고, 모델이 "찾을 수 없습니다"로 판단하도록 한다.
    relevant = [c for c in top if c["score"] >= SCORE_THRESHOLD]
    return relevant if relevant else top[:1]


# ---------- 생성 ----------
SYSTEM_PROMPT = (
    "당신은 NoteCook의 문서 분석 도우미입니다. "
    "아래 제공된 '참고 문맥'에 근거해서만 한국어로 답하세요. "
    "문맥에 없는 내용은 추측하지 말고 '제공된 문서에서 찾을 수 없습니다'라고 답하세요. "
    "답변에는 근거가 된 부분을 [번호] 형태로 인용하세요."
)


def _cited_numbers(text: str) -> set[int]:
    """답변 본문에서 [1], [1, 2] 형태로 실제 인용된 출처 번호를 추출."""
    nums = set()
    for grp in re.findall(r"\[([\d,\s]+)\]", text):
        for n in re.findall(r"\d+", grp):
            nums.add(int(n))
    return nums


def build_context(chunks: list[dict]) -> tuple[str, list[dict]]:
    """검색된 청크로 컨텍스트 문자열 + 출처 목록 생성."""
    lines = []
    sources = []
    for i, c in enumerate(chunks, start=1):
        loc = f" · p.{c['page']}" if c.get("page") else ""
        lines.append(f"[{i}] (출처: {c['doc_name']}{loc})\n{c['text']}")
        sources.append({"n": i, "name": c["doc_name"], "page": c.get("page"),
                        "doc_id": c.get("doc_id"), "score": round(c["score"], 3)})
    return "\n\n".join(lines), sources


def answer(question: str, doc_ids: list[str], model: str,
           embed_model: str = EMBED_MODEL, on_token=None) -> dict:
    """
    RAG 답변. {answer, sources} 반환.

    on_token(delta) 콜백을 주면 Ollama 를 스트리밍 모드로 호출하고
    생성되는 토큰 조각마다 콜백을 호출한다(타이핑 효과). 콜백이 없으면
    기존처럼 한 번에 받아 반환한다.
    """
    if not model:
        return {"error": "사용할 모델이 선택되지 않았습니다. 우측 상단에서 모델을 선택하세요."}
    chunks = search(question, doc_ids, embed_model=embed_model)
    if not chunks:
        return {"error": "검색할 문서가 없습니다. 먼저 문서를 추가하세요."}

    context, sources = build_context(chunks)
    user_msg = f"참고 문맥:\n{context}\n\n질문: {question}"

    stream = on_token is not None

    def _chat_once(think=False) -> str:
        # think=False: 문서 Q&A 에는 추론이 불필요하고, 추론이 토큰 예산을 다 써
        # 본문(content)을 못 내는 빈 응답의 원인이 되므로 끈다.
        payload = {
            "model": model,
            "messages": [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_msg},
            ],
            "stream": stream,
            "options": {"num_ctx": NUM_CTX},
        }
        if think is not None:
            payload["think"] = think
        r = requests.post(f"{OLLAMA_HOST}/api/chat", json=payload,
                          timeout=300, stream=stream)
        # 일부 모델은 think 파라미터 미지원 → 400. think 빼고 한 번 재시도.
        if r.status_code == 400 and think is not None:
            return _chat_once(think=None)
        r.raise_for_status()
        if not stream:
            body = r.json()
            return body.get("message", {}).get("content", "").strip()
        parts = []
        for line in r.iter_lines(decode_unicode=True):
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError:
                continue
            delta = (obj.get("message") or {}).get("content", "")
            if delta:
                parts.append(delta)
                try:
                    on_token(delta)
                except Exception:
                    pass
            if obj.get("done"):
                break
        return "".join(parts).strip()

    content = _chat_once()
    if not content:
        # 안전망: 그래도 빈 내용이면 모델 로드를 보장한 뒤 한 번 재시도.
        from backend import ollama as _ol
        _ol.load_model(model)   # 로드가 끝날 때까지 블록
        content = _chat_once()
    # 모델이 답변에서 실제 [n] 으로 인용한 출처만 남긴다.
    # 인용이 전혀 없으면(모델이 표기를 누락한 경우) 임계값을 통과한 검색 출처를 그대로 둔다.
    cited = _cited_numbers(content)
    if cited:
        sources = [s for s in sources if s["n"] in cited]
    # 중복 출처(같은 문서·페이지) 정리. 단, 같은 출처를 가리키는 인용 번호들은
    # nums 로 모아 둔다 (본문의 [n] ↔ 출처 칩 매핑을 프런트에서 복원하기 위함).
    by_key: dict = {}
    order = []
    for s in sources:
        key = (s["name"], s["page"])
        if key not in by_key:
            by_key[key] = {**s, "nums": []}
            order.append(key)
        by_key[key]["nums"].append(s["n"])
    uniq = [by_key[k] for k in order]
    return {"answer": content, "sources": uniq}
